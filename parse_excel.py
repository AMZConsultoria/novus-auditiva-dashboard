import openpyxl
import re
from datetime import datetime

MONTHS = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

def clean_value(val):
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    s = str(val).strip()
    if s in ('-', '—', 'R$ -', 'R$-'):
        return 0.0
    s = re.sub(r'R\$\s*', '', s)
    s = s.replace('.', '').replace(',', '.').strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {
        'ultima_atualizacao': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'empresa': 'Novus Auditiva — Juvevê',
        'ano': '2026',
        'mes_atual': 'MAI',
        'categorias': [],
        'kpis': {},
        'resumo_mensal': {},
    }

    ws = wb['Quadro Resumo']

    # Detect header row and month/total columns
    month_cols = {}
    total_col = None
    header_row_num = 3  # default

    for row in ws.iter_rows(min_row=1, max_row=8):
        for cell in row:
            v = str(cell.value).strip().upper() if cell.value else ''
            if v in MONTHS and v not in month_cols:
                month_cols[v] = cell.column
                header_row_num = cell.row
            if v in ('2026', 'TOTAL', '2026 TOTAL', 'ANO'):
                total_col = cell.column

    # Parse data rows
    cats = []
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row):
        label_cell = ws.cell(row=row[0].row, column=2)
        label = str(label_cell.value).strip() if label_cell.value else ''
        if not label:
            continue

        values = {}
        for m in MONTHS:
            col = month_cols.get(m)
            values[m] = clean_value(ws.cell(row=row[0].row, column=col).value) if col else 0.0

        total = clean_value(ws.cell(row=row[0].row, column=total_col).value) if total_col else sum(values.values())

        cats.append({'label': label, 'values': values, 'total': total})

    result['categorias'] = cats

    # Build monthly summary dict
    for m in MONTHS:
        result['resumo_mensal'][m] = {c['label']: c['values'].get(m, 0) for c in cats}

    # Detect most recent month with data
    rec = next((c for c in cats if 'RECEITA BRUTA' in c['label'].upper()), None)
    if rec:
        for m in reversed(MONTHS):
            if rec['values'].get(m, 0) != 0:
                result['mes_atual'] = m
                break

    # Build KPIs
    def find(keyword):
        return next((c for c in cats if keyword.lower() in c['label'].lower()), None)

    mes = result['mes_atual']
    rb  = find('RECEITA BRUTA')
    rl  = find('RECEITA LÍQUIDA')
    dep = find('(-) DESPESAS')
    res = find('RESULTADO MENSAL (CAIXA)')
    dis = find('DISTRIBUIÇÃO DE LUCROS')

    def v(cat, m=None):
        if not cat:
            return 0.0
        return cat['values'].get(m, 0.0) if m else cat.get('total', 0.0)

    result['kpis'] = {
        'receita_bruta_mes':    v(rb, mes),
        'receita_bruta_anual':  v(rb),
        'receita_liq_mes':      v(rl, mes),
        'receita_liq_anual':    v(rl),
        'despesas_mes':         abs(v(dep, mes)),
        'despesas_anual':       abs(v(dep)),
        'resultado_mes':        v(res, mes),
        'resultado_anual':      v(res),
        'distribuicao_mes':     abs(v(dis, mes)),
        'distribuicao_anual':   abs(v(dis)),
    }

    wb.close()
    return result
